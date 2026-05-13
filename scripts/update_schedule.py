#!/usr/bin/env python3
"""Update KU Shuttle schedule data from the transportation Excel file."""

from __future__ import annotations

import argparse
import json
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None


LOCATIONS = {
    "al_rawda": "Al Rawda",
    "kurh": "KURH",
    "umm_lulu": "Umm Lulu",
    "masdar": "Masdar",
    "main": "Main Campus",
    "san": "SAN Campus",
    "al_bahar": "Al Bahar Parking",
}

DORMS = ["al_rawda", "kurh", "umm_lulu", "masdar"]
MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return re.sub(r"[ \t]+", " ", str(value).replace("\r", "\n")).strip()


def parse_time_token(text: str) -> str | None:
    text = text.strip()
    if not text:
        return None
    if re.fullmatch(r"\d{1,2}:\d{2}", text):
        hour, minute = [int(part) for part in text.split(":")]
        return f"{hour:02d}:{minute:02d}"
    match = re.search(r"(\d{1,2})(?::(\d{2}))?\s*(AM|PM)", text, re.I)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or "00")
    suffix = match.group(3).upper()
    if suffix == "PM" and hour != 12:
        hour += 12
    if suffix == "AM" and hour == 12:
        hour = 0
    return f"{hour:02d}:{minute:02d}"


def parse_lines(value):
    if isinstance(value, time):
        return [{"time": value.strftime("%H:%M"), "scope": "all", "raw": ""}]

    text = clean(value)
    if not text:
        return []

    rows = []
    for line in [part.strip() for part in text.split("\n") if part.strip()]:
        parsed = parse_time_token(line)
        if not parsed:
            continue

        low = line.lower()
        scope = "all"
        if low.startswith("masdar"):
            scope = "masdar"
        elif low.startswith("other residences"):
            scope = "non_masdar"

        note = re.sub(r"\b\d{1,2}(?::\d{2})?\s*(?:AM|PM)\b", "", line, flags=re.I)
        note = note.replace("Masdar", "").replace("Other Residences & Main Campus", "Other residences")
        note = note.strip(" -")
        rows.append({"time": parsed, "scope": scope, "raw": note})
    return rows


def parse_period(label: str, sheet_name: str) -> dict:
    merged = clean(label) or sheet_name
    year_match = re.search(r"(20\d{2})", merged)
    year = int(year_match.group(1)) if year_match else 2026
    month_match = re.search(r"(Jan|January|Feb|February|Mar|March|Apr|April|May|Jun|June|Jul|July|Aug|August|Sep|September|Oct|October|Nov|November|Dec|December)", merged, re.I)
    month = MONTHS[month_match.group(1).lower()] if month_match else 5
    nums = [int(x) for x in re.findall(r"\b(\d{1,2})\b", merged)]
    day_numbers = [n for n in nums if 1 <= n <= 31 and n != year]
    if "to" in merged.lower() and len(day_numbers) >= 2:
      start_day, end_day = day_numbers[0], day_numbers[-1]
    else:
      start_day = end_day = day_numbers[0] if day_numbers else 1
    start = date(year, month, start_day)
    end = date(year, month, end_day)
    if start == end:
        label_short = start.strftime("%a %d %b")
    else:
        label_short = f"{start.strftime('%d %b')} - {end.strftime('%d %b')}"
    return {
        "id": start.isoformat() if start == end else f"{start.isoformat()}..{end.isoformat()}",
        "label": label_short,
        "fullLabel": merged.replace("\n", " "),
        "start": start.isoformat(),
        "end": end.isoformat(),
    }


def add_entries(routes, period_id, origins, destinations, entries, exam_label, exam_time):
    for entry in entries:
        scoped_origins = origins
        if entry["scope"] == "masdar":
            scoped_origins = [origin for origin in origins if origin == "masdar"]
        elif entry["scope"] == "non_masdar":
            scoped_origins = [origin for origin in origins if origin != "masdar"]

        note_bits = [exam_label]
        if exam_time:
            note_bits.append(clean(exam_time).replace("\n", " / "))
        if entry["raw"]:
            note_bits.append(entry["raw"])
        note = " - ".join(note_bits)

        for origin in scoped_origins:
            for destination in destinations:
                routes[period_id][f"{origin}>{destination}"].append({"t": entry["time"], "note": note})


def column_index(cell_ref: str) -> int:
    letters = "".join(char for char in cell_ref if char.isalpha())
    index = 0
    for char in letters:
        index = index * 26 + ord(char.upper()) - ord("A") + 1
    return index - 1


def excel_number(value: str):
    try:
        numeric = float(value)
    except ValueError:
        return value
    if 0 <= numeric < 1:
        minutes = round(numeric * 24 * 60)
        return time(minutes // 60, minutes % 60)
    return value


def read_xlsx_rows(path: Path):
    spreadsheet_ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(path) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("a:si", spreadsheet_ns):
                shared_strings.append("".join(node.text or "" for node in item.findall(".//a:t", spreadsheet_ns)))

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in relationships.findall("rel:Relationship", package_rel_ns)
        }

        for sheet in workbook.findall("a:sheets/a:sheet", spreadsheet_ns):
            sheet_name = sheet.attrib["name"]
            relationship_id = sheet.attrib[f"{{{office_rel_ns}}}id"]
            target = relationship_targets[relationship_id]
            sheet_path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
            sheet_root = ET.fromstring(archive.read(sheet_path))
            rows = []

            for row in sheet_root.findall("a:sheetData/a:row", spreadsheet_ns):
                values = []
                for cell in row.findall("a:c", spreadsheet_ns):
                    ref = cell.attrib.get("r", "")
                    index = column_index(ref)
                    while len(values) <= index:
                        values.append(None)

                    cell_type = cell.attrib.get("t")
                    raw_value = cell.find("a:v", spreadsheet_ns)
                    inline = cell.find("a:is", spreadsheet_ns)
                    value = None
                    if cell_type == "s" and raw_value is not None:
                        value = shared_strings[int(raw_value.text)]
                    elif cell_type == "inlineStr" and inline is not None:
                        value = "".join(node.text or "" for node in inline.findall(".//a:t", spreadsheet_ns))
                    elif raw_value is not None and raw_value.text is not None:
                        value = excel_number(raw_value.text)
                    values[index] = value

                rows.append(tuple(values))

            yield sheet_name, rows


def workbook_rows(path: Path):
    if load_workbook is not None:
        workbook = load_workbook(path, data_only=True)
        for worksheet in workbook.worksheets:
            yield worksheet.title, list(worksheet.iter_rows(values_only=True))
        return

    yield from read_xlsx_rows(path)


def parse_workbook(path: Path) -> dict:
    periods = []
    routes = defaultdict(lambda: defaultdict(list))

    for sheet_name, rows in workbook_rows(path):
        current_period = None

        for idx, row in enumerate(rows):
            section = clean(row[1] if len(row) > 1 else "")
            if section not in {"Main Campus Exam", "Between Main Campus and Al Bahar Parking", "SAN Campus Exam"}:
                continue

            first_data = idx + 2
            if section == "Main Campus Exam":
                outbound_origins, outbound_destinations = DORMS + ["san"], ["main"]
                return_origins, return_destinations = ["main"], DORMS + ["san"]
                label = "Main Campus exam"
            elif section == "Between Main Campus and Al Bahar Parking":
                outbound_origins, outbound_destinations = ["al_bahar"], ["main"]
                return_origins, return_destinations = ["main"], ["al_bahar"]
                label = "Al Bahar parking shuttle"
            else:
                outbound_origins, outbound_destinations = DORMS + ["main"], ["san"]
                return_origins, return_destinations = ["san"], DORMS + ["main"]
                label = "SAN Campus exam"

            row_i = first_data
            while row_i < len(rows):
                data = rows[row_i]
                if all(clean(cell) == "" for cell in data):
                    break

                possible_new_section = clean(data[1] if len(data) > 1 else "")
                if possible_new_section in {"Main Campus Exam", "Between Main Campus and Al Bahar Parking", "SAN Campus Exam"}:
                    break

                period_text = clean(data[1] if len(data) > 1 else "")
                if period_text:
                    current_period = parse_period(period_text, sheet_name)
                    if not any(period["id"] == current_period["id"] for period in periods):
                        periods.append(current_period)

                if current_period:
                    exam_time = data[2] if len(data) > 2 else ""
                    outbound = parse_lines(data[3] if len(data) > 3 else "")
                    inbound = parse_lines(data[4] if len(data) > 4 else "")
                    add_entries(routes, current_period["id"], outbound_origins, outbound_destinations, outbound, label, exam_time)
                    add_entries(routes, current_period["id"], return_origins, return_destinations, inbound, label, exam_time)

                row_i += 1

    clean_routes = {}
    for period_id, period_routes in routes.items():
        clean_routes[period_id] = {}
        for route_key, entries in period_routes.items():
            unique = {}
            for entry in entries:
                unique[(entry["t"], entry["note"])] = entry
            clean_routes[period_id][route_key] = sorted(unique.values(), key=lambda item: item["t"])

    return {
        "meta": {
            "title": "Final Exam Bus Schedule",
            "version": "Friday 08 May to Thursday 14 May 2026",
            "sourceFile": path.name,
            "updatedAt": datetime.now().strftime("%Y-%m-%d"),
        },
        "locations": LOCATIONS,
        "order": ["al_rawda", "kurh", "umm_lulu", "masdar", "main", "san", "al_bahar"],
        "periods": sorted(periods, key=lambda period: period["start"]),
        "routes": clean_routes,
    }


def update_html(html_path: Path, data: dict) -> None:
    text = html_path.read_text(encoding="utf-8")
    blob = "const APP_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";"
    pattern = re.compile(r"const APP_DATA = .*?;\n\s*/\* END_SCHEDULE_DATA \*/", re.S)
    replacement = blob + "\n    /* END_SCHEDULE_DATA */"
    updated, count = pattern.subn(replacement, text)
    if count != 1:
        raise RuntimeError("Could not find APP_DATA block in index.html")
    html_path.write_text(updated, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate KU Shuttle schedule data from an Excel file.")
    parser.add_argument("source", type=Path)
    parser.add_argument("--data", type=Path, default=Path("data/schedule-current.json"))
    parser.add_argument("--html", type=Path, default=Path("index.html"))
    args = parser.parse_args()

    data = parse_workbook(args.source)
    args.data.parent.mkdir(parents=True, exist_ok=True)
    args.data.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.html.exists():
        update_html(args.html, data)
    print(f"Updated {args.data} from {args.source.name}")


if __name__ == "__main__":
    main()
